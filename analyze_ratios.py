import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook

def update_financial_ratios(file_path="raw_data.xlsx"):
    try:
        # --- Step 1: Read Raw Financial Data from Sheet1 ---
        raw_df = pd.read_excel(file_path, sheet_name="Sheet1", engine="openpyxl")

        # --- Step 2: Calculate Financial Ratios ---
        df = pd.DataFrame()
        df["Company"] = raw_df["Company"]
        df["Net Profit Margin"] = raw_df["Net Income"] / raw_df["Revenue"]
        df["Return on Assets (ROA)"] = raw_df["Net Income"] / raw_df["Total Assets"]
        df["Current Ratio"] = raw_df["Current Assets"] / raw_df["Current Liabilities"]
        df["Quick Ratio"] = (raw_df["Current Assets"] - raw_df.get("Inventory", 0)) / raw_df["Current Liabilities"]
        df["Debt to Equity"] = raw_df["Total Debt"] / raw_df["Shareholder Equity"]
        df["Interest Coverage Ratio"] = raw_df["Net Income"] / raw_df["Interest Expense"]

        # --- Step 3: Load Workbook ---
        wb = load_workbook(file_path)

        # If "Ratio" sheet does not exist, create it
        if "Ratio" not in wb.sheetnames:
            ws = wb.create_sheet("Ratio")
        else:
            ws = wb["Ratio"]

        # --- Step 4: Overwrite the Ratio Sheet (preserving charts if present) ---
        rows = dataframe_to_rows(df, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # --- Step 5: Save Workbook ---
        wb.save(file_path)
        print("✅ Ratios updated based on data from Sheet1!")

    except FileNotFoundError:
        print("❌ Excel file not found.")
    except KeyError as e:
        print(f"❌ Missing column in Sheet1: {e}")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")

# Run
if __name__ == "__main__":
    update_financial_ratios()
